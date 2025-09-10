import pandas as pd
import os
import subprocess
import shutil
import tempfile
import re
import json
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
# work with different sheets for submissions and results
# section wise in the tabs in sheet
# TODO work in onedrive sheets, get details from tab -> sections {choose the tab name}

# --- CONFIGURATION ---
# IMPORTANT: Update these values in env to match your setup

# 1. Path to your Excel file in your local OneDrive folder
# STUDENT_SUBMISSIONS = "C:/Users/YourUsername/OneDrive/StudentProjects.xlsx" # Example for Windows
STUDENT_SUBMISSIONS = os.path.expanduser(os.getenv("STUDENT_SUBMISSIONS")) # Example for Mac/Linux
STUDENT_RESULTS = os.path.expanduser(os.getenv("STUDENT_RESULTS"))
STUDENT_NAME_COLUMN = "Student Name"
PROGRAM_TIMEOUT = 15 # A shorter timeout is fine for simple programs

# NEW: Input to provide to the Java program's standard input.
# Use '\n' to simulate the user pressing the Enter key.
#TODO set the input based on the question for different assignments
# PROGRAM_INPUT = "4003600000000014\n0\n" # Provides 25.0 for the first prompt, 10.0 for the second.

with open("workshop_inputs.json", "r", encoding="utf-8") as f:
    WORKSHOP_TESTS = json.load(f)
# --- END OF CONFIGURATION ---

def find_file(directory, filename):
    for root, dirs, files in os.walk(directory):
        if filename in files:
            return os.path.join(root, filename)
    return None

def add_dropdown_to_status_column(file_path, sheet_name, workshop_number):
    """
    Adds a dropdown list to the specified status column in the given sheet.
    Dropdown options: ["⛔Absent", "❌Incomplete", "⚠️Partial Complete", "✅Complete"]
    """
    try:
        status_column_index = workshop_number + 1  # Column B is 2, C is 3, etc.
        status_column_letter = get_column_letter(status_column_index)
        # Load the workbook and sheet
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in workbook.")
            return
        sheet = workbook[sheet_name]

        # Create a data validation object for dropdown
        dropdown = DataValidation(type="list", formula1='"⛔Absent,❌Incomplete,⚠️Partial Complete,✅Complete"', allow_blank=True)
        sheet.add_data_validation(dropdown)

        # Apply the dropdown to the entire column
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
            cell = sheet[f"{status_column_letter}{row}"]
            dropdown.add(cell)

        # Save the workbook
        workbook.save(file_path)
        print(f"Dropdown added successfully to column '{status_column_letter}' in sheet '{sheet_name}'.")
    except Exception as e:
        print(f"Error adding dropdown: {e}")
        
def run_command(command, working_dir, input_data=None):
    """
    Runs a command, optionally providing input data to its stdin.
    Returns its status and output.
    """
    try:
        result = subprocess.run(
            command,
            cwd=working_dir,
            check=False,
            capture_output=True,
            text=True,
            timeout=PROGRAM_TIMEOUT,
            input=input_data # Pass input to the process
        )
        return result
    except FileNotFoundError:
        print(f"Error: Command '{command[0]}' not found. Is it in your system's PATH?")
        return None
    except subprocess.TimeoutExpired:
        print(f"  Error: Command timed out after {PROGRAM_TIMEOUT} seconds.")
        return "Timeout"

def detect_main_class(java_files):
    """Detect the main class with package if present."""
    for file in java_files:
        with open(file, "r", encoding="utf-8") as f:
            content = f.read()

            if re.search(r'public\s+static\s+void\s+main\s*\(\s*String\s*\[\]\s*\w+\)', content):
                class_name = os.path.splitext(os.path.basename(file))[0]

                # Check for package declaration
                package_match = re.search(r'package\s+([\w\.]+);', content)
                if package_match:
                    return package_match.group(1) + "." + class_name
                else:
                    return class_name
    return None

def run_tests(clone_path, main_class, tests):
    """
    Runs multiple test cases for a compiled Java program.
    Each test contains 'input' and optionally 'expected'.
    Returns final status and detailed results.
    """
    results_summary = []
    passed_count = 0

    for i, test in enumerate(tests, start=1):
        print(f"  Running test case {i}...")
        input_data = test["input"]
        expected = test.get("expected")

        run_command_list = ["java", "-cp", clone_path, main_class]
        run_result = run_command(run_command_list, clone_path, input_data)

        if run_result == "Timeout":
            results_summary.append(f"Test {i}: Timeout ❌")
            print(results_summary[-1])
            continue
        if run_result is None or run_result.returncode != 0:
            error_message = run_result.stderr if run_result else "Java command failed"
            results_summary.append(f"Test {i}: Runtime Error ❌ -> {error_message}")
            print(results_summary[-1])
            continue

        program_output = (run_result.stdout + run_result.stderr).strip()

        if expected:  # Compare expected vs actual
            if expected not in program_output:
                results_summary.append(
                    f"Test {i}: Failed ❌\n"
                    f"Input:\n{input_data}"
                    f"Expected: \n{expected}\n"
                    f"Got:\n{program_output}\n"
                )
            else:
                results_summary.append(
                    f"Test {i}: Passed ✅\n"
                    f"Input:\n{input_data}"
                    f"Expected: \n{expected}\n"
                    f"Got:\n{program_output}\n"
                )
                passed_count += 1
        else:  # No expected → just record output
            results_summary.append(
                f"Test {i}: Output captured"
                f"Input:\n{input_data}"
                f"Expected:\n{expected}\n"
                f"Got:\n{program_output}\n"
            )
            passed_count += 1  # treat as pass if no expectation

        print(results_summary[-1])  # print last test result
        
    total_tests = len(tests)
    
    if passed_count == total_tests:
        return "✅Complete", f"Summary: {passed_count}/{total_tests} tests passed"
    elif passed_count > 0:
        return "Partial Complete", f"Summary: {passed_count}/{total_tests} tests passed"
    else:
        return "Runtime Error", f"Summary: {passed_count}/{total_tests} tests passed"

    
def process_student_repo(repo_url, tests):
    """
    Clones, compiles, and runs a student's Java project.
    Returns a status string and any relevant error messages.
    """
    if not repo_url or pd.isna(repo_url):
        return "Absent", "No repository URL provided."

    with tempfile.TemporaryDirectory() as temp_dir:
        clone_path = os.path.join(temp_dir, "repo")
        print(f"  Cloning {repo_url}...")

        clone_command = ["git", "clone", repo_url, clone_path]
        clone_result = run_command(clone_command, temp_dir)

        if clone_result == "Timeout":
            return "Git Clone Error", "Git clone timed out."
        if clone_result is None or clone_result.returncode != 0:
            error_message = clone_result.stderr if clone_result else "Git command failed."
            return "Git Clone Error", f"Failed to clone repo.\n{error_message}"
        
        # Look for .java files
        java_files = []
        for root, dirs, files in os.walk(clone_path):
            for fname in files:
                if fname.endswith(".java"):
                    java_files.append(os.path.join(root, fname))
        if not java_files:
            return "Incomplete", "No .java files found in the repository."

        main_class = detect_main_class(java_files)
        print(f"  Detected main class: {main_class}")
        if main_class is None:
            return "Incomplete", "Could not find a class with a main method."

        # compile all java files from repo root so package structure is preserved
        print(f"  Compiling Java files ({len(java_files)} files)...")
        relative_java_files = [os.path.relpath(f, clone_path) for f in java_files]
        compile_command = ["javac", "-d", clone_path] + relative_java_files
        compile_result = run_command(compile_command, clone_path)

        if compile_result == "Timeout":
            return "Compile Error", "Compiler timed out."
        if compile_result is None or compile_result.returncode != 0:
            error_message = compile_result.stderr if compile_result else "Javac command failed."
            return "Compile Error", f"Code did not compile.\n{error_message}"

        # --- run test cases ---
        return run_tests(clone_path, main_class, tests)

def main():
    """Main function to drive the script.
     Ask user which workshop column to use 
    (columns named "Workshop 1 Repo URL" .. "Workshop 11 Repo URL")
     """
    print("--- Starting Student Project Grader ---")  
    
    # Ask section
    section = input("Enter section number: ").strip().upper()
    INPUT_SHEET_NAME = f"L2C{section}"
    OUTPUT_SHEET_NAME = f"L2C{section}"
    print(f"Working on section: {INPUT_SHEET_NAME}")
    
    while True:
        try:
            workshop = int(input("Enter workshop number (1-11): "))
            if 1 <= workshop <= 11:
                break
        except ValueError:
            pass
        print("Please enter a number between 1 and 11.")
        
    # Validate if the workshop exists in the JSON
    if str(workshop) not in WORKSHOP_TESTS:
        print(f"Error: Workshop {workshop} is not defined in the JSON file.")
        return
    
    REPO_URL_COLUMN = f"Workshop {workshop} Repo URL"
    print(f"Using repository column: '{REPO_URL_COLUMN}'")

    if not os.path.exists(STUDENT_SUBMISSIONS):
        print(f"Error: The file '{STUDENT_SUBMISSIONS}' was not found.")
        return
    
    # get the tests for the workshop
    tests = WORKSHOP_TESTS.get(str(workshop), {}).get("tests", [])   
    
    try:
        df = pd.read_excel(STUDENT_SUBMISSIONS, sheet_name=INPUT_SHEET_NAME)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    results = []
    for index, row in df.iterrows():
        student_name = row[STUDENT_NAME_COLUMN]
        repo_url = row.get(REPO_URL_COLUMN)
        print(f"\nProcessing {student_name}...")
        
        status, details = process_student_repo(repo_url, tests)
        
        final_status = {
            "Absent": "⛔Absent",
            "Git Clone Error": "❌Incomplete",
            "Compile Error": "❌Incomplete",
            "Runtime Error": "⚠️Partial Complete",
            "Partial Complete": "⚠️Partial Complete",
            "✅Complete": "✅Complete"
        }.get(status.strip(), "Unknown Error")

        print(f"  Status: {status} -> {final_status}")
        
        results.append({
            STUDENT_NAME_COLUMN: student_name,
            f"Workshop {workshop} Status": final_status,
            f"Workshop {workshop} Details": details
        })

    results_df = pd.DataFrame(results)
        

    print(f"\nWriting results to sheet '{OUTPUT_SHEET_NAME}' in '{STUDENT_RESULTS}'...")
    ''' FIXME if there are no students name but has results then it will fail, gives -> Error writing to Excel file: cannot reindex on an axis with duplicate labels
    if there is no data in the workshop with column name format, it start writing from the first column with the workshop column name
    it is possible to check or set status of workshops of later weeks even if the previous weeks statuses are not there
    Works properly if the student name column has student names'''
    try:
        status_col = f"Workshop {workshop} Status"
        # details_col = f"Workshop {workshop} Details"
        # Read existing Results sheet if present
        if os.path.exists(STUDENT_RESULTS):
            try:
                existing = pd.read_excel(STUDENT_RESULTS, sheet_name=OUTPUT_SHEET_NAME)
            except Exception:
                existing = pd.DataFrame()
        else:
            existing = pd.DataFrame()

        new = results_df.copy()  # results_df contains STUDENT_NAME_COLUMN and the two workshop columns
        print(f"Evaluated results: \n",new)
        # If an existing Results sheet has a student name column, merge by student name
        if not existing.empty and STUDENT_NAME_COLUMN in existing.columns:
            existing = existing.set_index(STUDENT_NAME_COLUMN)
            new = new.set_index(STUDENT_NAME_COLUMN)

            all_index = existing.index.union(new.index)
            existing = existing.reindex(all_index)

            # Update/insert the two workshop columns from the new results (aligns by student)
            existing[status_col] = new[status_col].reindex(all_index)
            # existing[details_col] = new[details_col].reindex(all_index)

            merged = existing.reset_index()
        else:
            # No existing results to merge with — use the new results as-is
            merged = new.reset_index() if STUDENT_NAME_COLUMN in new.index.names else new

        # Write back just the Results sheet (replace it) while preserving other sheets in the workbook
        mode = 'a' if os.path.exists(STUDENT_RESULTS) else 'w'
        with pd.ExcelWriter(STUDENT_RESULTS, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            merged.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME, index=False)

        add_dropdown_to_status_column(STUDENT_RESULTS, OUTPUT_SHEET_NAME, workshop)  # Assuming column B for status
        print("--- Script finished successfully! ---")
    except Exception as e:
        print(f"\nError writing to Excel file: {e}")

if __name__ == "__main__":
    main()